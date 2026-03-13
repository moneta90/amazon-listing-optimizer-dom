#!/usr/bin/env python3

import argparse
import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit("Missing dependency: openpyxl") from exc

try:
    import xlrd
except ImportError:
    extra_path = Path("/tmp/codex_xlrd")
    if extra_path.exists():
        sys.path.insert(0, str(extra_path))
        import xlrd
    else:
        raise SystemExit("Missing dependency: xlrd")


def normalize_filename(name):
    return re.sub(r" \(\d+\)(?=\.)", "", name)


def detect_domain(filename):
    return re.sub(r"^de_|_browse_tree_guide.*$", "", filename)


def slugify(text):
    text = (text or "").strip().lower()
    replacements = {
        "ä": "ae",
        "ö": "oe",
        "ü": "ue",
        "ß": "ss",
        "&": " und ",
    }
    for source, target in replacements.items():
        text = text.replace(source, target)
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-")


def to_node_id(value):
    if value is None or value == "":
        return None
    if isinstance(value, float):
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    text = str(value).strip()
    if not text:
        return None
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def normalize_path(path):
    return str(path).strip().replace("/", " > ")


def iter_xls_rows(path):
    book = xlrd.open_workbook(path)
    category_sheet = book.sheet_by_name(book.sheet_names()[1])
    refinements_sheet = book.sheet_by_name("Refinements")
    categories = (category_sheet.row_values(index) for index in range(1, category_sheet.nrows))
    refinements = (refinements_sheet.row_values(index) for index in range(1, refinements_sheet.nrows))
    return categories, refinements


def iter_xlsx_rows(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    category_sheet = workbook[workbook.sheetnames[1]]
    refinements_sheet = workbook["Refinements"]
    categories = category_sheet.iter_rows(min_row=2, values_only=True)
    refinements = refinements_sheet.iter_rows(min_row=2, values_only=True)
    return categories, refinements


def build_dataset(input_dir):
    input_path = Path(input_dir)
    files = sorted(p for p in input_path.iterdir() if p.suffix in {".xls", ".xlsx"})

    selected_files = []
    seen_names = set()
    for path in files:
        normalized_name = normalize_filename(path.name)
        if normalized_name in seen_names:
            continue
        seen_names.add(normalized_name)
        selected_files.append(path)

    categories_by_id = {}
    category_attrs = defaultdict(lambda: {"node_id": None, "item_type": None, "path": None, "attrs": [], "domain": None})
    attribute_counter = Counter()

    for path in selected_files:
        domain = detect_domain(path.name)
        if path.suffix == ".xls":
            category_rows, refinement_rows = iter_xls_rows(path)
        else:
            category_rows, refinement_rows = iter_xlsx_rows(path)

        local_paths = {}
        for row in category_rows:
            if not row:
                continue
            node_id = to_node_id(row[0])
            node_path = row[1] if len(row) > 1 else None
            if not node_id or not node_path:
                continue

            normalized_path = normalize_path(node_path)
            leaf = normalized_path.split(" > ")[-1]
            item_type = slugify(leaf)
            category_id = f"{domain}:{node_id}"
            local_paths[node_id] = (category_id, normalized_path, item_type)

            category = {
                "id": category_id,
                "node_id": node_id,
                "path": normalized_path,
                "item_type": item_type,
                "attr_count": 0,
                "domain": domain,
            }
            previous = categories_by_id.get(category_id)
            if previous is None or len(normalized_path) > len(previous["path"]):
                categories_by_id[category_id] = category

        for row in refinement_rows:
            if not row:
                continue
            node_id = to_node_id(row[0])
            if not node_id:
                continue

            attribute = ""
            if len(row) > 3 and row[3]:
                attribute = str(row[3]).strip()

            if node_id in local_paths:
                category_id, normalized_path, item_type = local_paths[node_id]
                category_attrs[category_id]["node_id"] = node_id
                category_attrs[category_id]["path"] = normalized_path
                category_attrs[category_id]["item_type"] = item_type
                category_attrs[category_id]["domain"] = domain
                if attribute:
                    category_attrs[category_id]["attrs"].append(attribute)
                    attribute_counter[attribute] += 1

    for category_id, category in categories_by_id.items():
        attr_entry = category_attrs.get(category_id)
        if attr_entry:
            category["attr_count"] = len(attr_entry["attrs"])

    serialized_attrs = {}
    for category_id, entry in category_attrs.items():
        if not entry["path"]:
            continue
        serialized_attrs[category_id] = {
            "node_id": entry["node_id"],
            "item_type": entry["item_type"],
            "path": entry["path"],
            "attrs": entry["attrs"],
            "domain": entry["domain"],
        }

    categories = sorted(categories_by_id.values(), key=lambda item: (item["domain"], item["path"]))
    top_attributes = [name for name, _ in attribute_counter.most_common(30)]

    return {
        "categories": categories,
        "category_attrs": serialized_attrs,
        "values": [],
        "top_attributes": top_attributes,
    }


def main():
    parser = argparse.ArgumentParser(description="Merge Amazon BTG spreadsheets into a single JSON dataset.")
    parser.add_argument("--input-dir", default="Browse Tree Guide - BTG - Kategorie")
    parser.add_argument("--output", default="public/btg-data.json")
    args = parser.parse_args()

    dataset = build_dataset(args.input_dir)
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(dataset, ensure_ascii=False), encoding="utf-8")

    print(
        json.dumps(
            {
                "files": args.input_dir,
                "categories": len(dataset["categories"]),
                "category_attrs": len(dataset["category_attrs"]),
                "top_attributes": dataset["top_attributes"][:10],
                "output": str(output_path),
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
